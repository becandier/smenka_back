# tests/test_penalties.py
"""Фича fines: шаблоны штрафов, назначение/снятие/правка, учёт в payroll, shifts.is_deleted."""

import uuid
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.app.core.security import hash_password
from src.app.models.member_rate import OrganizationMemberRate, RateType
from src.app.models.organization import MemberRole, Organization, OrganizationMember
from src.app.models.penalty import Penalty
from src.app.models.shift import Shift, ShiftStatus
from src.app.models.user import User

RATE_EFF_JAN = datetime(2026, 1, 1, 0, 0, 0, tzinfo=UTC)


# --- helpers -----------------------------------------------------------------
async def _make_finished_shift(
    db_session: AsyncSession,
    user_id: uuid.UUID,
    org_id: uuid.UUID,
    started_at: datetime,
    finished_at: datetime,
) -> Shift:
    shift = Shift(
        user_id=user_id,
        organization_id=org_id,
        started_at=started_at,
        finished_at=finished_at,
        status=ShiftStatus.finished,
    )
    db_session.add(shift)
    await db_session.commit()
    return shift


async def _make_rate(
    db_session: AsyncSession,
    member_id: uuid.UUID,
    amount: int,
    rate_type: RateType = RateType.hourly,
) -> OrganizationMemberRate:
    rate = OrganizationMemberRate(
        member_id=member_id,
        rate_amount_minor=amount,
        rate_type=rate_type,
        currency="RUB",
        effective_from=RATE_EFF_JAN,
    )
    db_session.add(rate)
    await db_session.commit()
    return rate


def _data(resp: Any) -> Any:
    return resp.json()["data"]


def _err(resp: Any) -> str:
    return resp.json()["error"]["code"]


async def _create_template(
    client: AsyncClient,
    headers: dict[str, str],
    org_id: uuid.UUID,
    reason: str = "Опоздание",
    amount_minor: int = 50000,
) -> dict[str, Any]:
    resp = await client.post(
        f"/api/v1/organizations/{org_id}/penalty-templates",
        headers=headers,
        json={"reason": reason, "amount_minor": amount_minor},
    )
    assert resp.status_code == 201, resp.text
    return _data(resp)


async def _create_penalty(
    client: AsyncClient,
    headers: dict[str, str],
    org_id: uuid.UUID,
    **body: Any,
) -> Any:
    return await client.post(
        f"/api/v1/organizations/{org_id}/penalties",
        headers=headers,
        json=body,
    )


# --- fixtures ----------------------------------------------------------------
@pytest.fixture
async def owner(db_session: AsyncSession) -> User:
    user = User(
        id=uuid.uuid4(),
        email="owner@example.com",
        password_hash=hash_password("Test1234"),
        name="Owner",
        is_verified=True,
    )
    db_session.add(user)
    await db_session.commit()
    return user


@pytest.fixture
async def owner_headers(owner: User, client: AsyncClient) -> dict[str, str]:
    resp = await client.post(
        "/api/v1/auth/login",
        json={"email": "owner@example.com", "password": "Test1234"},
    )
    token = resp.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
async def admin_user(db_session: AsyncSession) -> User:
    user = User(
        id=uuid.uuid4(),
        email="orgadmin@example.com",
        password_hash=hash_password("Test1234"),
        name="Org Admin",
        is_verified=True,
    )
    db_session.add(user)
    await db_session.commit()
    return user


@pytest.fixture
async def admin_headers(admin_user: User, client: AsyncClient) -> dict[str, str]:
    resp = await client.post(
        "/api/v1/auth/login",
        json={"email": "orgadmin@example.com", "password": "Test1234"},
    )
    token = resp.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
async def emp2_user(db_session: AsyncSession) -> User:
    user = User(
        id=uuid.uuid4(),
        email="emp2@example.com",
        password_hash=hash_password("Test1234"),
        name="Employee Two",
        is_verified=True,
    )
    db_session.add(user)
    await db_session.commit()
    return user


@pytest.fixture
async def emp2_headers(emp2_user: User, client: AsyncClient) -> dict[str, str]:
    resp = await client.post(
        "/api/v1/auth/login",
        json={"email": "emp2@example.com", "password": "Test1234"},
    )
    token = resp.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
async def org(db_session: AsyncSession, owner: User) -> Organization:
    organization = Organization(name="Fines Org", owner_id=owner.id)
    db_session.add(organization)
    await db_session.commit()
    return organization


@pytest.fixture
async def employee_member(
    db_session: AsyncSession,
    org: Organization,
    verified_user: User,
) -> OrganizationMember:
    member = OrganizationMember(
        organization_id=org.id,
        user_id=verified_user.id,
        role=MemberRole.employee,
    )
    db_session.add(member)
    await db_session.commit()
    return member


@pytest.fixture
async def admin_member(
    db_session: AsyncSession,
    org: Organization,
    admin_user: User,
) -> OrganizationMember:
    member = OrganizationMember(
        organization_id=org.id,
        user_id=admin_user.id,
        role=MemberRole.admin,
    )
    db_session.add(member)
    await db_session.commit()
    return member


@pytest.fixture
async def emp2_member(
    db_session: AsyncSession,
    org: Organization,
    emp2_user: User,
) -> OrganizationMember:
    member = OrganizationMember(
        organization_id=org.id,
        user_id=emp2_user.id,
        role=MemberRole.employee,
    )
    db_session.add(member)
    await db_session.commit()
    return member


# --- Шаблоны -----------------------------------------------------------------
async def test_create_template(client, owner_headers, org):
    data = await _create_template(client, owner_headers, org.id, "Опоздание", 50000)
    assert data["reason"] == "Опоздание"
    assert data["amount_minor"] == 50000
    assert data["currency"] == "RUB"
    assert data["id"]
    assert data["created_at"]
    assert data["updated_at"]


async def test_create_template_validation(client, owner_headers, org):
    bad_amount = await client.post(
        f"/api/v1/organizations/{org.id}/penalty-templates",
        headers=owner_headers,
        json={"reason": "X", "amount_minor": 0},
    )
    assert bad_amount.status_code == 422
    bad_reason = await client.post(
        f"/api/v1/organizations/{org.id}/penalty-templates",
        headers=owner_headers,
        json={"reason": "", "amount_minor": 100},
    )
    assert bad_reason.status_code == 422


async def test_list_templates_excludes_deleted(client, owner_headers, org):
    t1 = await _create_template(client, owner_headers, org.id, "A", 1000)
    await _create_template(client, owner_headers, org.id, "B", 2000)
    resp = await client.get(
        f"/api/v1/organizations/{org.id}/penalty-templates", headers=owner_headers
    )
    assert resp.status_code == 200
    items = _data(resp)["items"]
    assert {i["reason"] for i in items} == {"A", "B"}

    # soft-delete t1
    dele = await client.delete(
        f"/api/v1/organizations/{org.id}/penalty-templates/{t1['id']}", headers=owner_headers
    )
    assert dele.status_code == 200
    assert _data(dele)["deleted"] is True
    resp2 = await client.get(
        f"/api/v1/organizations/{org.id}/penalty-templates", headers=owner_headers
    )
    items2 = _data(resp2)["items"]
    assert {i["reason"] for i in items2} == {"B"}


async def test_update_template(client, owner_headers, org):
    t = await _create_template(client, owner_headers, org.id, "Опоздание", 50000)
    resp = await client.patch(
        f"/api/v1/organizations/{org.id}/penalty-templates/{t['id']}",
        headers=owner_headers,
        json={"amount_minor": 60000, "reason": "Опоздание > 15 мин"},
    )
    assert resp.status_code == 200
    data = _data(resp)
    assert data["amount_minor"] == 60000
    assert data["reason"] == "Опоздание > 15 мин"


async def test_delete_template_twice_404(client, owner_headers, org):
    t = await _create_template(client, owner_headers, org.id)
    first = await client.delete(
        f"/api/v1/organizations/{org.id}/penalty-templates/{t['id']}", headers=owner_headers
    )
    assert first.status_code == 200
    second = await client.delete(
        f"/api/v1/organizations/{org.id}/penalty-templates/{t['id']}", headers=owner_headers
    )
    assert second.status_code == 404
    assert _err(second) == "PENALTY_TEMPLATE_NOT_FOUND"


async def test_template_rbac(client, owner_headers, admin_headers, auth_headers, org,
                             admin_member, employee_member):
    # admin (member) — можно
    resp_admin = await client.post(
        f"/api/v1/organizations/{org.id}/penalty-templates",
        headers=admin_headers,
        json={"reason": "Admin", "amount_minor": 100},
    )
    assert resp_admin.status_code == 201
    # employee — нельзя
    resp_emp = await client.post(
        f"/api/v1/organizations/{org.id}/penalty-templates",
        headers=auth_headers,
        json={"reason": "Emp", "amount_minor": 100},
    )
    assert resp_emp.status_code == 403
    assert _err(resp_emp) == "FORBIDDEN"


# --- Штрафы: создание --------------------------------------------------------
async def test_create_custom_penalty(client, owner_headers, owner, org, employee_member,
                                      verified_user):
    resp = await _create_penalty(
        client, owner_headers, org.id,
        member_id=str(employee_member.id),
        reason="Брак",
        amount_minor=30000,
        occurred_at="2026-06-15T08:00:00Z",
    )
    assert resp.status_code == 201, resp.text
    data = _data(resp)
    assert data["member_id"] == str(employee_member.id)
    assert data["user_id"] == str(verified_user.id)
    assert data["user_name"] == "Test User"
    assert data["reason"] == "Брак"
    assert data["amount_minor"] == 30000
    assert data["currency"] == "RUB"
    assert data["template_id"] is None
    assert data["shift_id"] is None
    assert data["created_by_user_id"] == str(owner.id)


async def test_penalty_snapshot_from_template(client, owner_headers, org, employee_member):
    t = await _create_template(client, owner_headers, org.id, "Опоздание", 50000)
    resp = await _create_penalty(
        client, owner_headers, org.id,
        member_id=str(employee_member.id),
        template_id=t["id"],
        occurred_at="2026-06-15T08:00:00Z",
    )
    assert resp.status_code == 201, resp.text
    data = _data(resp)
    assert data["reason"] == "Опоздание"
    assert data["amount_minor"] == 50000
    assert data["template_id"] == t["id"]


async def test_penalty_override_template(client, owner_headers, org, employee_member):
    t = await _create_template(client, owner_headers, org.id, "Опоздание", 50000)
    resp = await _create_penalty(
        client, owner_headers, org.id,
        member_id=str(employee_member.id),
        template_id=t["id"],
        amount_minor=60000,
        occurred_at="2026-06-15T08:00:00Z",
    )
    data = _data(resp)
    assert data["reason"] == "Опоздание"  # из шаблона
    assert data["amount_minor"] == 60000  # переопределение


async def test_snapshot_independent_of_template_edit(client, owner_headers, org, employee_member):
    t = await _create_template(client, owner_headers, org.id, "Опоздание", 50000)
    p = _data(
        await _create_penalty(
            client, owner_headers, org.id,
            member_id=str(employee_member.id),
            template_id=t["id"],
            occurred_at="2026-06-15T08:00:00Z",
        )
    )
    # правим шаблон
    await client.patch(
        f"/api/v1/organizations/{org.id}/penalty-templates/{t['id']}",
        headers=owner_headers,
        json={"amount_minor": 99999, "reason": "Изменено"},
    )
    got = _data(
        await client.get(
            f"/api/v1/organizations/{org.id}/penalties/{p['id']}", headers=owner_headers
        )
    )
    assert got["amount_minor"] == 50000
    assert got["reason"] == "Опоздание"


async def test_occurred_at_from_shift(client, owner_headers, db_session, org, employee_member,
                                      verified_user):
    started = datetime(2026, 6, 10, 9, 0, tzinfo=UTC)
    shift = await _make_finished_shift(
        db_session, verified_user.id, org.id, started, datetime(2026, 6, 10, 17, 0, tzinfo=UTC)
    )
    resp = await _create_penalty(
        client, owner_headers, org.id,
        member_id=str(employee_member.id),
        reason="Опоздание",
        amount_minor=10000,
        shift_id=str(shift.id),
    )
    assert resp.status_code == 201, resp.text
    data = _data(resp)
    assert data["shift_id"] == str(shift.id)
    assert datetime.fromisoformat(data["occurred_at"]) == started


async def test_occurred_at_explicit_with_shift(client, owner_headers, db_session, org,
                                                employee_member, verified_user):
    shift = await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 9, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 17, 0, tzinfo=UTC),
    )
    resp = await _create_penalty(
        client, owner_headers, org.id,
        member_id=str(employee_member.id),
        reason="X", amount_minor=10000,
        shift_id=str(shift.id),
        occurred_at="2026-06-11T00:00:00Z",
    )
    data = _data(resp)
    assert datetime.fromisoformat(data["occurred_at"]) == datetime(2026, 6, 11, 0, 0, tzinfo=UTC)


async def test_occurred_at_required_without_shift(client, owner_headers, org, employee_member):
    resp = await _create_penalty(
        client, owner_headers, org.id,
        member_id=str(employee_member.id),
        reason="X", amount_minor=10000,
    )
    assert resp.status_code == 422
    assert _err(resp) == "VALIDATION_ERROR"


async def test_shift_of_another_member_404(client, owner_headers, db_session, org,
                                           employee_member, emp2_member, emp2_user):
    other_shift = await _make_finished_shift(
        db_session, emp2_user.id, org.id,
        datetime(2026, 6, 10, 9, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 17, 0, tzinfo=UTC),
    )
    resp = await _create_penalty(
        client, owner_headers, org.id,
        member_id=str(employee_member.id),
        reason="X", amount_minor=10000,
        shift_id=str(other_shift.id),
    )
    assert resp.status_code == 404
    assert _err(resp) == "SHIFT_NOT_FOUND"


async def test_member_not_found(client, owner_headers, org):
    resp = await _create_penalty(
        client, owner_headers, org.id,
        member_id=str(uuid.uuid4()),
        reason="X", amount_minor=10000,
        occurred_at="2026-06-15T08:00:00Z",
    )
    assert resp.status_code == 404
    assert _err(resp) == "MEMBER_NOT_FOUND"


async def test_deleted_template_not_usable(client, owner_headers, org, employee_member):
    t = await _create_template(client, owner_headers, org.id)
    await client.delete(
        f"/api/v1/organizations/{org.id}/penalty-templates/{t['id']}", headers=owner_headers
    )
    resp = await _create_penalty(
        client, owner_headers, org.id,
        member_id=str(employee_member.id),
        template_id=t["id"],
        occurred_at="2026-06-15T08:00:00Z",
    )
    assert resp.status_code == 404
    assert _err(resp) == "PENALTY_TEMPLATE_NOT_FOUND"


# --- Штрафы: список/деталь/правка/снятие ------------------------------------
async def test_list_penalties_filters_and_pagination(
    client, owner_headers, org, employee_member, emp2_member
):
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="a", amount_minor=1000, occurred_at="2026-06-10T00:00:00Z")
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="b", amount_minor=2000, occurred_at="2026-06-20T00:00:00Z")
    await _create_penalty(client, owner_headers, org.id, member_id=str(emp2_member.id),
                          reason="c", amount_minor=3000, occurred_at="2026-06-15T00:00:00Z")

    all_resp = await client.get(
        f"/api/v1/organizations/{org.id}/penalties", headers=owner_headers
    )
    body = _data(all_resp)
    assert body["total"] == 3
    # сортировка occurred_at DESC
    occ = [datetime.fromisoformat(i["occurred_at"]) for i in body["items"]]
    assert occ == sorted(occ, reverse=True)

    by_member = _data(
        await client.get(
            f"/api/v1/organizations/{org.id}/penalties?member_id={employee_member.id}",
            headers=owner_headers,
        )
    )
    assert by_member["total"] == 2

    by_date = _data(
        await client.get(
            f"/api/v1/organizations/{org.id}/penalties"
            f"?date_from=2026-06-15T00:00:00Z&date_to=2026-06-20T00:00:00Z",
            headers=owner_headers,
        )
    )
    assert by_date["total"] == 2  # 06-15 и 06-20 включительно

    page = _data(
        await client.get(
            f"/api/v1/organizations/{org.id}/penalties?limit=1&offset=0", headers=owner_headers
        )
    )
    assert page["total"] == 3
    assert len(page["items"]) == 1


async def test_get_and_soft_delete_penalty(
    client, owner_headers, admin_headers, db_session, org, employee_member,
    admin_member, admin_user
):
    p = _data(
        await _create_penalty(
            client, owner_headers, org.id, member_id=str(employee_member.id),
            reason="X", amount_minor=10000, occurred_at="2026-06-15T08:00:00Z",
        )
    )
    # снимает admin (не автор-owner)
    dele = await client.delete(
        f"/api/v1/organizations/{org.id}/penalties/{p['id']}", headers=admin_headers
    )
    assert dele.status_code == 200
    assert _data(dele)["deleted"] is True

    # деталь снятого → 404
    got = await client.get(
        f"/api/v1/organizations/{org.id}/penalties/{p['id']}", headers=owner_headers
    )
    assert got.status_code == 404
    assert _err(got) == "PENALTY_NOT_FOUND"

    # повторный DELETE → 404
    again = await client.delete(
        f"/api/v1/organizations/{org.id}/penalties/{p['id']}", headers=owner_headers
    )
    assert again.status_code == 404
    assert _err(again) == "PENALTY_NOT_FOUND"

    # зафиксирован кто и когда снял
    row = (
        await db_session.execute(select(Penalty).where(Penalty.id == uuid.UUID(p["id"])))
    ).scalar_one()
    assert row.is_deleted is True
    assert row.deleted_by_user_id == admin_user.id
    assert row.deleted_at is not None


async def test_update_penalty(client, owner_headers, db_session, org, employee_member,
                              verified_user):
    shift = await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 9, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 17, 0, tzinfo=UTC),
    )
    p = _data(
        await _create_penalty(
            client, owner_headers, org.id, member_id=str(employee_member.id),
            reason="X", amount_minor=10000, occurred_at="2026-06-15T08:00:00Z",
        )
    )
    resp = await client.patch(
        f"/api/v1/organizations/{org.id}/penalties/{p['id']}",
        headers=owner_headers,
        json={"amount_minor": 40000, "reason": "Y", "shift_id": str(shift.id), "comment": "note"},
    )
    assert resp.status_code == 200
    data = _data(resp)
    assert data["amount_minor"] == 40000
    assert data["reason"] == "Y"
    assert data["shift_id"] == str(shift.id)
    assert data["comment"] == "note"

    # обнулить shift_id
    resp2 = await client.patch(
        f"/api/v1/organizations/{org.id}/penalties/{p['id']}",
        headers=owner_headers,
        json={"shift_id": None},
    )
    assert _data(resp2)["shift_id"] is None


async def test_update_penalty_shift_other_member_404(
    client, owner_headers, db_session, org, employee_member, emp2_member, emp2_user
):
    other_shift = await _make_finished_shift(
        db_session, emp2_user.id, org.id,
        datetime(2026, 6, 10, 9, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 17, 0, tzinfo=UTC),
    )
    p = _data(
        await _create_penalty(
            client, owner_headers, org.id, member_id=str(employee_member.id),
            reason="X", amount_minor=10000, occurred_at="2026-06-15T08:00:00Z",
        )
    )
    resp = await client.patch(
        f"/api/v1/organizations/{org.id}/penalties/{p['id']}",
        headers=owner_headers,
        json={"shift_id": str(other_shift.id)},
    )
    assert resp.status_code == 404
    assert _err(resp) == "SHIFT_NOT_FOUND"


# --- my-penalties ------------------------------------------------------------
async def test_my_penalties_employee_sees_own(
    client, owner_headers, auth_headers, org, employee_member
):
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=10000, occurred_at="2026-06-15T08:00:00Z")
    resp = await client.get(f"/api/v1/organizations/{org.id}/my-penalties", headers=auth_headers)
    assert resp.status_code == 200
    body = _data(resp)
    assert body["total"] == 1
    assert body["items"][0]["amount_minor"] == 10000
    assert "user_id" not in body["items"][0]  # employee-facing форма


async def test_my_penalties_owner_forbidden(client, owner_headers, org):
    resp = await client.get(f"/api/v1/organizations/{org.id}/my-penalties", headers=owner_headers)
    assert resp.status_code == 403
    assert _err(resp) == "FORBIDDEN"


async def test_my_penalties_isolated_between_employees(
    client, owner_headers, emp2_headers, org, employee_member, emp2_member
):
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=10000, occurred_at="2026-06-15T08:00:00Z")
    resp = await client.get(f"/api/v1/organizations/{org.id}/my-penalties", headers=emp2_headers)
    assert _data(resp)["total"] == 0


async def test_penalty_rbac_employee_denied(client, auth_headers, org, employee_member):
    create = await _create_penalty(
        client, auth_headers, org.id, member_id=str(employee_member.id),
        reason="X", amount_minor=10000, occurred_at="2026-06-15T08:00:00Z",
    )
    assert create.status_code == 403
    listing = await client.get(f"/api/v1/organizations/{org.id}/penalties", headers=auth_headers)
    assert listing.status_code == 403


# --- интеграция в payroll ----------------------------------------------------
async def _payroll(client, headers, org_id, **params):
    qs = "&".join(f"{k}={v}" for k, v in params.items())
    url = f"/api/v1/organizations/{org_id}/payroll"
    if qs:
        url += f"?{qs}"
    return await client.get(url, headers=headers)


async def test_payroll_net(client, owner_headers, db_session, org, employee_member, verified_user):
    await _make_rate(db_session, employee_member.id, 18000)  # 180₽/час
    await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),  # 2ч → 36000
    )
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=10000, occurred_at="2026-06-10T12:00:00Z")
    resp = await _payroll(client, owner_headers, org.id)
    data = _data(resp)
    item = data["items"][0]
    assert item["gross_amount_minor"] == 36000
    assert item["penalty_amount_minor"] == 10000
    assert item["penalties_count"] == 1
    assert item["net_amount_minor"] == 26000
    assert data["totals"]["penalty_amount_minor"] == 10000
    assert data["totals"]["net_amount_minor"] == 26000


async def test_payroll_net_negative(client, owner_headers, db_session, org, employee_member,
                                    verified_user):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),  # 36000
    )
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=50000, occurred_at="2026-06-10T12:00:00Z")
    item = _data(await _payroll(client, owner_headers, org.id))["items"][0]
    assert item["net_amount_minor"] == -14000


async def test_payroll_penalty_without_rate(client, owner_headers, db_session, org,
                                            employee_member, verified_user):
    # смена есть, ставки нет → gross 0, но штраф учитывается
    await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=5000, occurred_at="2026-06-10T12:00:00Z")
    item = _data(await _payroll(client, owner_headers, org.id))["items"][0]
    assert item["gross_amount_minor"] == 0
    assert item["net_amount_minor"] == -5000


async def test_payroll_penalty_only_member_in_items(
    client, owner_headers, org, employee_member, verified_user
):
    # ни одной смены, только штраф → сотрудник всё равно в items
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=5000, occurred_at="2026-06-10T12:00:00Z")
    data = _data(await _payroll(client, owner_headers, org.id))
    item = next(i for i in data["items"] if i["user_id"] == str(verified_user.id))
    assert item["gross_amount_minor"] == 0
    assert item["shifts_count"] == 0
    assert item["net_amount_minor"] == -5000


async def test_payroll_include_penalties_false(
    client, owner_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=10000, occurred_at="2026-06-10T12:00:00Z")
    item = _data(
        await _payroll(client, owner_headers, org.id, include_penalties="false")
    )["items"][0]
    assert item["penalty_amount_minor"] == 0
    assert item["penalties_count"] == 0
    assert item["net_amount_minor"] == 36000


async def test_payroll_include_penalties_false_drops_penalty_only_member(
    client, owner_headers, org, employee_member, verified_user
):
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=5000, occurred_at="2026-06-10T12:00:00Z")
    data = _data(await _payroll(client, owner_headers, org.id, include_penalties="false"))
    assert all(i["user_id"] != str(verified_user.id) for i in data["items"])


async def test_payroll_penalty_period_filter(
    client, owner_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    # штраф вне периода
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=10000, occurred_at="2026-07-01T00:00:00Z")
    item = _data(
        await _payroll(
            client, owner_headers, org.id,
            date_from="2026-06-01T00:00:00Z", date_to="2026-06-30T23:59:59Z",
        )
    )["items"][0]
    assert item["penalty_amount_minor"] == 0
    assert item["net_amount_minor"] == 36000


async def test_my_earnings_includes_penalty(
    client, owner_headers, auth_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=10000, occurred_at="2026-06-10T12:00:00Z")
    resp = await client.get(
        f"/api/v1/organizations/{org.id}/my-earnings", headers=auth_headers
    )
    data = _data(resp)
    assert data["gross_amount_minor"] == 36000
    assert data["penalty_amount_minor"] == 10000
    assert data["penalties_count"] == 1
    assert data["net_amount_minor"] == 26000


async def test_payroll_export_penalty_columns(
    client, owner_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=10000, occurred_at="2026-06-10T12:00:00Z")
    resp = await client.get(
        f"/api/v1/organizations/{org.id}/payroll/export", headers=owner_headers
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    summary = wb["Сводка"]
    rows = list(summary.iter_rows(values_only=True))
    header = next(r for r in rows if r and r[0] == "Сотрудник")
    assert "Штраф, ₽" in header
    assert "К выплате, ₽" in header
    p_idx = header.index("Штраф, ₽")
    n_idx = header.index("К выплате, ₽")
    emp_row = next(r for r in rows if r and r[0] == "Test User")
    assert emp_row[p_idx] == 100.0  # 10000 коп.
    assert emp_row[n_idx] == 260.0  # 26000 коп.


# --- shifts.is_deleted -------------------------------------------------------
async def test_soft_deleted_shift_excluded_everywhere(
    client, owner_headers, auth_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    shift = await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    # до удаления — смена видна в payroll
    pre = _data(await _payroll(client, owner_headers, org.id))
    assert pre["items"][0]["gross_amount_minor"] == 36000

    # soft-delete смены напрямую (эндпоинта удаления нет — заготовка)
    shift.is_deleted = True
    await db_session.commit()

    # payroll: смена исключена → нет items
    post = _data(await _payroll(client, owner_headers, org.id))
    assert post["items"] == []

    # org-shifts список: пусто
    org_list = _data(
        await client.get(f"/api/v1/organizations/{org.id}/shifts", headers=owner_headers)
    )
    assert org_list["total"] == 0

    # org-shift деталь: 404
    detail = await client.get(
        f"/api/v1/organizations/{org.id}/shifts/{shift.id}", headers=owner_headers
    )
    assert detail.status_code == 404

    # личный список смен сотрудника: пусто
    my_shifts = _data(await client.get("/api/v1/shifts", headers=auth_headers))
    assert my_shifts["total"] == 0


async def test_soft_deleted_shift_excluded_from_admin_user_detail(
    client, super_admin_headers, db_session, org, employee_member, verified_user
):
    shift = await _make_finished_shift(
        db_session, verified_user.id, org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    pre = _data(
        await client.get(
            f"/api/v1/admin/users/{verified_user.id}", headers=super_admin_headers
        )
    )
    assert pre["shifts_count"] == 1

    shift.is_deleted = True
    await db_session.commit()

    post = _data(
        await client.get(
            f"/api/v1/admin/users/{verified_user.id}", headers=super_admin_headers
        )
    )
    assert post["shifts_count"] == 0


async def test_payroll_only_missing_rate_keeps_penalty_only_member(
    client, owner_headers, org, employee_member, verified_user
):
    # penalty-only сотрудник (без смен) не исчезает под only_missing_rate — штраф не теряется
    await _create_penalty(client, owner_headers, org.id, member_id=str(employee_member.id),
                          reason="X", amount_minor=5000, occurred_at="2026-06-10T12:00:00Z")
    data = _data(await _payroll(client, owner_headers, org.id, only_missing_rate="true"))
    item = next(i for i in data["items"] if i["user_id"] == str(verified_user.id))
    assert item["penalty_amount_minor"] == 5000
    assert item["net_amount_minor"] == -5000
    assert data["totals"]["penalty_amount_minor"] == 5000
