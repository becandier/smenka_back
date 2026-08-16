# tests/test_adjustments.py
"""Фича manual_time_entry (B): ручные начисления (payroll_adjustments) + payroll."""

import uuid
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.app.core.security import hash_password
from src.app.models.audit_log import AuditLog
from src.app.models.member_rate import OrganizationMemberRate, RateType
from src.app.models.notification import Notification
from src.app.models.organization import MemberRole, Organization, OrganizationMember
from src.app.models.shift import Shift, ShiftStatus
from src.app.models.user import User

RATE_EFF_JAN = datetime(2026, 1, 1, tzinfo=UTC)


def _data(resp: Any) -> Any:
    return resp.json()["data"]


def _err(resp: Any) -> str:
    return resp.json()["error"]["code"]


# --- fixtures ------------------------------------------------------------------
@pytest.fixture
async def owner(db_session: AsyncSession) -> User:
    user = User(
        id=uuid.uuid4(),
        email="adj_owner@example.com",
        password_hash=hash_password("Test1234"),
        name="Owner",
        is_verified=True,
    )
    db_session.add(user)
    await db_session.commit()
    return user


@pytest.fixture
async def owner_headers(owner: User, client: AsyncClient) -> dict[str, str]:
    resp = await client.post(
        "/api/v1/auth/login",
        json={"email": "adj_owner@example.com", "password": "Test1234"},
    )
    token = resp.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
async def emp2_user(db_session: AsyncSession) -> User:
    user = User(
        id=uuid.uuid4(),
        email="adj_emp2@example.com",
        password_hash=hash_password("Test1234"),
        name="Employee Two",
        is_verified=True,
    )
    db_session.add(user)
    await db_session.commit()
    return user


@pytest.fixture
async def emp2_headers(emp2_user: User, client: AsyncClient) -> dict[str, str]:
    resp = await client.post(
        "/api/v1/auth/login",
        json={"email": "adj_emp2@example.com", "password": "Test1234"},
    )
    token = resp.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
async def org(db_session: AsyncSession, owner: User) -> Organization:
    organization = Organization(name="Adjustments Org", owner_id=owner.id)
    db_session.add(organization)
    await db_session.commit()
    return organization


@pytest.fixture
async def employee_member(
    db_session: AsyncSession,
    org: Organization,
    verified_user: User,
) -> OrganizationMember:
    member = OrganizationMember(
        organization_id=org.id,
        user_id=verified_user.id,
        role=MemberRole.employee,
    )
    db_session.add(member)
    await db_session.commit()
    return member


@pytest.fixture
async def emp2_member(
    db_session: AsyncSession,
    org: Organization,
    emp2_user: User,
) -> OrganizationMember:
    member = OrganizationMember(
        organization_id=org.id,
        user_id=emp2_user.id,
        role=MemberRole.employee,
    )
    db_session.add(member)
    await db_session.commit()
    return member


async def _make_finished_shift(
    db_session: AsyncSession,
    user_id: uuid.UUID,
    org_id: uuid.UUID,
    started_at: datetime,
    finished_at: datetime,
) -> Shift:
    shift = Shift(
        user_id=user_id,
        organization_id=org_id,
        started_at=started_at,
        finished_at=finished_at,
        status=ShiftStatus.finished,
    )
    db_session.add(shift)
    await db_session.commit()
    return shift


async def _make_rate(
    db_session: AsyncSession,
    member_id: uuid.UUID,
    amount: int,
    rate_type: RateType = RateType.hourly,
) -> OrganizationMemberRate:
    rate = OrganizationMemberRate(
        member_id=member_id,
        rate_amount_minor=amount,
        rate_type=rate_type,
        currency="RUB",
        effective_from=RATE_EFF_JAN,
    )
    db_session.add(rate)
    await db_session.commit()
    return rate


async def _create_adjustment(
    client: AsyncClient, headers: dict[str, str], org_id: uuid.UUID, **body: Any
) -> Any:
    return await client.post(
        f"/api/v1/organizations/{org_id}/adjustments", headers=headers, json=body
    )


# --- B1: создать начисление ------------------------------------------------------
async def test_create_adjustment_positive(
    client, owner_headers, owner, org, employee_member, verified_user, db_session
):
    resp = await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=300000,
        reason="Доплата за 2 ч (забыл отметиться)",
        occurred_at="2026-06-15T00:00:00Z",
        comment="Согласовано с бригадиром",
    )
    assert resp.status_code == 201, resp.text
    data = _data(resp)
    assert data["member_id"] == str(employee_member.id)
    assert data["user_id"] == str(verified_user.id)
    assert data["user_name"] == "Test User"
    assert data["amount_minor"] == 300000
    assert data["currency"] == "RUB"
    assert data["reason"] == "Доплата за 2 ч (забыл отметиться)"
    assert data["comment"] == "Согласовано с бригадиром"
    assert data["created_by_user_id"] == str(owner.id)
    assert data["created_by_name"] == "Owner"

    audit = (
        (await db_session.execute(select(AuditLog).where(AuditLog.action == "adjustment.create")))
        .scalars()
        .all()
    )
    assert len(audit) == 1
    assert audit[0].actor_user_id == owner.id

    notif = (
        (
            await db_session.execute(
                select(Notification).where(Notification.user_id == verified_user.id)
            )
        )
        .scalars()
        .all()
    )
    assert len(notif) == 1
    assert notif[0].type == "payroll_adjustment_changed"
    assert notif[0].payload["action"] == "created"
    assert notif[0].payload["amount_minor"] == 300000


async def test_create_adjustment_negative(client, owner_headers, org, employee_member):
    resp = await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=-50000,
        reason="Разовое удержание",
        occurred_at="2026-06-15T00:00:00Z",
    )
    assert resp.status_code == 201, resp.text
    assert _data(resp)["amount_minor"] == -50000


async def test_create_adjustment_amount_zero_422(client, owner_headers, org, employee_member):
    resp = await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=0,
        reason="X",
        occurred_at="2026-06-15T00:00:00Z",
    )
    assert resp.status_code == 422


async def test_create_adjustment_occurred_at_from_shift(
    client, owner_headers, db_session, org, employee_member, verified_user
):
    started = datetime(2026, 6, 10, 9, 0, tzinfo=UTC)
    shift = await _make_finished_shift(
        db_session, verified_user.id, org.id, started, datetime(2026, 6, 10, 17, 0, tzinfo=UTC)
    )
    resp = await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=10000,
        reason="X",
        shift_id=str(shift.id),
    )
    assert resp.status_code == 201, resp.text
    assert datetime.fromisoformat(_data(resp)["occurred_at"]) == started


async def test_create_adjustment_occurred_at_required_without_shift_422(
    client, owner_headers, org, employee_member
):
    resp = await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=10000,
        reason="X",
    )
    assert resp.status_code == 422
    assert _err(resp) == "VALIDATION_ERROR"


async def test_create_adjustment_member_not_found_404(client, owner_headers, org):
    resp = await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(uuid.uuid4()),
        amount_minor=10000,
        reason="X",
        occurred_at="2026-06-15T00:00:00Z",
    )
    assert resp.status_code == 404
    assert _err(resp) == "MEMBER_NOT_FOUND"


async def test_create_adjustment_shift_of_other_member_404(
    client, owner_headers, db_session, org, employee_member, emp2_member, emp2_user
):
    other_shift = await _make_finished_shift(
        db_session,
        emp2_user.id,
        org.id,
        datetime(2026, 6, 10, 9, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 17, 0, tzinfo=UTC),
    )
    resp = await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=10000,
        reason="X",
        shift_id=str(other_shift.id),
    )
    assert resp.status_code == 404
    assert _err(resp) == "SHIFT_NOT_FOUND"


async def test_create_adjustment_forbidden_employee_and_super_admin(
    client, auth_headers, super_admin_headers, org, employee_member
):
    for headers in (auth_headers, super_admin_headers):
        resp = await _create_adjustment(
            client,
            headers,
            org.id,
            member_id=str(employee_member.id),
            amount_minor=10000,
            reason="X",
            occurred_at="2026-06-15T00:00:00Z",
        )
        assert resp.status_code == 403
        assert _err(resp) == "FORBIDDEN"


# --- B2: список ------------------------------------------------------------------
async def test_list_adjustments_filters_and_pagination(
    client, owner_headers, org, employee_member, emp2_member
):
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=1000,
        reason="a",
        occurred_at="2026-06-10T00:00:00Z",
    )
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=-2000,
        reason="b",
        occurred_at="2026-06-20T00:00:00Z",
    )
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(emp2_member.id),
        amount_minor=3000,
        reason="c",
        occurred_at="2026-06-15T00:00:00Z",
    )

    all_resp = _data(
        await client.get(f"/api/v1/organizations/{org.id}/adjustments", headers=owner_headers)
    )
    assert all_resp["total"] == 3
    occ = [datetime.fromisoformat(i["occurred_at"]) for i in all_resp["items"]]
    assert occ == sorted(occ, reverse=True)

    by_member = _data(
        await client.get(
            f"/api/v1/organizations/{org.id}/adjustments?member_id={employee_member.id}",
            headers=owner_headers,
        )
    )
    assert by_member["total"] == 2

    by_date = _data(
        await client.get(
            f"/api/v1/organizations/{org.id}/adjustments"
            f"?date_from=2026-06-15T00:00:00Z&date_to=2026-06-20T00:00:00Z",
            headers=owner_headers,
        )
    )
    assert by_date["total"] == 2

    page = _data(
        await client.get(
            f"/api/v1/organizations/{org.id}/adjustments?limit=1&offset=0", headers=owner_headers
        )
    )
    assert page["total"] == 3
    assert len(page["items"]) == 1


# --- B3: правка --------------------------------------------------------------------
async def test_update_adjustment(client, owner_headers, db_session, org, employee_member):
    a = _data(
        await _create_adjustment(
            client,
            owner_headers,
            org.id,
            member_id=str(employee_member.id),
            amount_minor=10000,
            reason="X",
            occurred_at="2026-06-15T00:00:00Z",
        )
    )
    resp = await client.patch(
        f"/api/v1/organizations/{org.id}/adjustments/{a['id']}",
        headers=owner_headers,
        json={"amount_minor": -40000, "reason": "Y", "comment": "note"},
    )
    assert resp.status_code == 200, resp.text
    data = _data(resp)
    assert data["amount_minor"] == -40000
    assert data["reason"] == "Y"
    assert data["comment"] == "note"

    audit = (
        (await db_session.execute(select(AuditLog).where(AuditLog.action == "adjustment.update")))
        .scalars()
        .all()
    )
    assert len(audit) == 1


async def test_update_adjustment_amount_zero_422(client, owner_headers, org, employee_member):
    a = _data(
        await _create_adjustment(
            client,
            owner_headers,
            org.id,
            member_id=str(employee_member.id),
            amount_minor=10000,
            reason="X",
            occurred_at="2026-06-15T00:00:00Z",
        )
    )
    resp = await client.patch(
        f"/api/v1/organizations/{org.id}/adjustments/{a['id']}",
        headers=owner_headers,
        json={"amount_minor": 0},
    )
    assert resp.status_code == 422


async def test_update_adjustment_shift_other_member_404(
    client, owner_headers, db_session, org, employee_member, emp2_member, emp2_user
):
    other_shift = await _make_finished_shift(
        db_session,
        emp2_user.id,
        org.id,
        datetime(2026, 6, 10, 9, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 17, 0, tzinfo=UTC),
    )
    a = _data(
        await _create_adjustment(
            client,
            owner_headers,
            org.id,
            member_id=str(employee_member.id),
            amount_minor=10000,
            reason="X",
            occurred_at="2026-06-15T00:00:00Z",
        )
    )
    resp = await client.patch(
        f"/api/v1/organizations/{org.id}/adjustments/{a['id']}",
        headers=owner_headers,
        json={"shift_id": str(other_shift.id)},
    )
    assert resp.status_code == 404
    assert _err(resp) == "SHIFT_NOT_FOUND"


async def test_update_adjustment_not_found_404(client, owner_headers, org):
    resp = await client.patch(
        f"/api/v1/organizations/{org.id}/adjustments/{uuid.uuid4()}",
        headers=owner_headers,
        json={"reason": "Y"},
    )
    assert resp.status_code == 404
    assert _err(resp) == "ADJUSTMENT_NOT_FOUND"


# --- B4: отмена -----------------------------------------------------------------
async def test_delete_adjustment_soft_delete(
    client, owner_headers, owner, db_session, org, employee_member, verified_user
):
    a = _data(
        await _create_adjustment(
            client,
            owner_headers,
            org.id,
            member_id=str(employee_member.id),
            amount_minor=10000,
            reason="X",
            occurred_at="2026-06-15T00:00:00Z",
        )
    )
    resp = await client.delete(
        f"/api/v1/organizations/{org.id}/adjustments/{a['id']}", headers=owner_headers
    )
    assert resp.status_code == 200
    assert _data(resp)["deleted"] is True

    got = await client.get(
        f"/api/v1/organizations/{org.id}/adjustments?member_id={employee_member.id}",
        headers=owner_headers,
    )
    assert _data(got)["total"] == 0

    notif = (
        (
            await db_session.execute(
                select(Notification).where(
                    Notification.user_id == verified_user.id,
                    Notification.type == "payroll_adjustment_changed",
                )
            )
        )
        .scalars()
        .all()
    )
    assert any(n.payload["action"] == "deleted" for n in notif)


async def test_delete_adjustment_twice_404(client, owner_headers, org, employee_member):
    a = _data(
        await _create_adjustment(
            client,
            owner_headers,
            org.id,
            member_id=str(employee_member.id),
            amount_minor=10000,
            reason="X",
            occurred_at="2026-06-15T00:00:00Z",
        )
    )
    first = await client.delete(
        f"/api/v1/organizations/{org.id}/adjustments/{a['id']}", headers=owner_headers
    )
    assert first.status_code == 200
    second = await client.delete(
        f"/api/v1/organizations/{org.id}/adjustments/{a['id']}", headers=owner_headers
    )
    assert second.status_code == 404
    assert _err(second) == "ADJUSTMENT_NOT_FOUND"


async def test_adjustment_include_deleted_and_restore_cycle(
    client, owner_headers, owner, db_session, org, employee_member, verified_user
):
    a = _data(
        await _create_adjustment(
            client,
            owner_headers,
            org.id,
            member_id=str(employee_member.id),
            amount_minor=10000,
            reason="X",
            occurred_at="2026-06-15T00:00:00Z",
        )
    )
    await client.delete(
        f"/api/v1/organizations/{org.id}/adjustments/{a['id']}", headers=owner_headers
    )

    hidden = _data(
        await client.get(f"/api/v1/organizations/{org.id}/adjustments", headers=owner_headers)
    )
    assert hidden["total"] == 0

    shown = _data(
        await client.get(
            f"/api/v1/organizations/{org.id}/adjustments",
            headers=owner_headers,
            params={"include_deleted": "true"},
        )
    )
    assert shown["total"] == 1
    assert shown["items"][0]["is_deleted"] is True

    restored = await client.post(
        f"/api/v1/organizations/{org.id}/adjustments/{a['id']}/restore", headers=owner_headers
    )
    assert restored.status_code == 200
    data = _data(restored)
    assert data["is_deleted"] is False
    assert data["deleted_at"] is None

    listed_again = _data(
        await client.get(f"/api/v1/organizations/{org.id}/adjustments", headers=owner_headers)
    )
    assert listed_again["total"] == 1

    audit = (
        (await db_session.execute(select(AuditLog).where(AuditLog.action == "adjustment.restore")))
        .scalars()
        .all()
    )
    assert len(audit) == 1

    notif = (
        (
            await db_session.execute(
                select(Notification).where(
                    Notification.user_id == verified_user.id,
                    Notification.type == "payroll_adjustment_changed",
                )
            )
        )
        .scalars()
        .all()
    )
    assert any(n.payload["action"] == "restored" for n in notif)


async def test_restore_adjustment_not_deleted_409(client, owner_headers, org, employee_member):
    a = _data(
        await _create_adjustment(
            client,
            owner_headers,
            org.id,
            member_id=str(employee_member.id),
            amount_minor=10000,
            reason="X",
            occurred_at="2026-06-15T00:00:00Z",
        )
    )
    resp = await client.post(
        f"/api/v1/organizations/{org.id}/adjustments/{a['id']}/restore", headers=owner_headers
    )
    assert resp.status_code == 409
    assert _err(resp) == "ADJUSTMENT_NOT_DELETED"


async def test_adjustment_rbac_employee_denied(client, auth_headers, org, employee_member):
    create = await _create_adjustment(
        client,
        auth_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=10000,
        reason="X",
        occurred_at="2026-06-15T00:00:00Z",
    )
    assert create.status_code == 403
    listing = await client.get(f"/api/v1/organizations/{org.id}/adjustments", headers=auth_headers)
    assert listing.status_code == 403


# --- B5: my-adjustments -----------------------------------------------------------
async def test_my_adjustments_employee_sees_own(
    client, owner_headers, auth_headers, org, employee_member
):
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=10000,
        reason="X",
        occurred_at="2026-06-15T00:00:00Z",
    )
    resp = await client.get(f"/api/v1/organizations/{org.id}/my-adjustments", headers=auth_headers)
    assert resp.status_code == 200
    body = _data(resp)
    assert body["total"] == 1
    assert body["items"][0]["amount_minor"] == 10000
    assert "user_id" not in body["items"][0]


async def test_my_adjustments_owner_forbidden(client, owner_headers, org):
    resp = await client.get(
        f"/api/v1/organizations/{org.id}/my-adjustments", headers=owner_headers
    )
    assert resp.status_code == 403
    assert _err(resp) == "FORBIDDEN"


async def test_my_adjustments_isolated_between_employees(
    client, owner_headers, emp2_headers, org, employee_member, emp2_member
):
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=10000,
        reason="X",
        occurred_at="2026-06-15T00:00:00Z",
    )
    resp = await client.get(f"/api/v1/organizations/{org.id}/my-adjustments", headers=emp2_headers)
    assert _data(resp)["total"] == 0


# --- интеграция в payroll ----------------------------------------------------
async def _payroll(client, headers, org_id, **params):
    qs = "&".join(f"{k}={v}" for k, v in params.items())
    url = f"/api/v1/organizations/{org_id}/payroll"
    if qs:
        url += f"?{qs}"
    return await client.get(url, headers=headers)


async def test_payroll_net_with_positive_adjustment(
    client, owner_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session,
        verified_user.id,
        org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=15000,
        reason="Премия",
        occurred_at="2026-06-10T12:00:00Z",
    )
    item = _data(await _payroll(client, owner_headers, org.id))["items"][0]
    assert item["gross_amount_minor"] == 36000
    assert item["adjustment_amount_minor"] == 15000
    assert item["adjustments_count"] == 1
    assert item["net_amount_minor"] == 51000


async def test_payroll_net_with_negative_adjustment(
    client, owner_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session,
        verified_user.id,
        org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=-50000,
        reason="Удержание",
        occurred_at="2026-06-10T12:00:00Z",
    )
    item = _data(await _payroll(client, owner_headers, org.id))["items"][0]
    assert item["adjustment_amount_minor"] == -50000
    assert item["net_amount_minor"] == -14000


async def test_payroll_penalty_and_adjustment_combined(
    client, owner_headers, db_session, org, employee_member, verified_user
):
    from src.app.models.organization import OrganizationMember  # noqa: F401

    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session,
        verified_user.id,
        org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await client.post(
        f"/api/v1/organizations/{org.id}/penalties",
        headers=owner_headers,
        json={
            "member_id": str(employee_member.id),
            "reason": "Опоздание",
            "amount_minor": 10000,
            "occurred_at": "2026-06-10T12:00:00Z",
        },
    )
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=20000,
        reason="Премия",
        occurred_at="2026-06-10T12:00:00Z",
    )
    item = _data(await _payroll(client, owner_headers, org.id))["items"][0]
    # 36000 (gross) - 10000 (penalty) + 20000 (adjustment) = 46000
    assert item["net_amount_minor"] == 46000


async def test_payroll_adjustment_only_member_in_items(
    client, owner_headers, org, employee_member, verified_user
):
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=5000,
        reason="X",
        occurred_at="2026-06-10T12:00:00Z",
    )
    data = _data(await _payroll(client, owner_headers, org.id))
    item = next(i for i in data["items"] if i["user_id"] == str(verified_user.id))
    assert item["gross_amount_minor"] == 0
    assert item["shifts_count"] == 0
    assert item["net_amount_minor"] == 5000


async def test_payroll_include_adjustments_false(
    client, owner_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session,
        verified_user.id,
        org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=15000,
        reason="Премия",
        occurred_at="2026-06-10T12:00:00Z",
    )
    item = _data(await _payroll(client, owner_headers, org.id, include_adjustments="false"))[
        "items"
    ][0]
    assert item["adjustment_amount_minor"] == 0
    assert item["adjustments_count"] == 0
    assert item["net_amount_minor"] == 36000


async def test_payroll_adjustment_period_filter(
    client, owner_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session,
        verified_user.id,
        org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=15000,
        reason="Премия за июль",
        occurred_at="2026-07-01T00:00:00Z",
    )
    item = _data(
        await _payroll(
            client,
            owner_headers,
            org.id,
            date_from="2026-06-01T00:00:00Z",
            date_to="2026-06-30T23:59:59Z",
        )
    )["items"][0]
    assert item["adjustment_amount_minor"] == 0
    assert item["net_amount_minor"] == 36000


async def test_my_earnings_includes_adjustment(
    client, owner_headers, auth_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session,
        verified_user.id,
        org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=15000,
        reason="Премия",
        occurred_at="2026-06-10T12:00:00Z",
    )
    resp = await client.get(f"/api/v1/organizations/{org.id}/my-earnings", headers=auth_headers)
    data = _data(resp)
    assert data["gross_amount_minor"] == 36000
    assert data["adjustment_amount_minor"] == 15000
    assert data["adjustments_count"] == 1
    assert data["net_amount_minor"] == 51000


async def test_payroll_export_adjustment_column(
    client, owner_headers, db_session, org, employee_member, verified_user
):
    await _make_rate(db_session, employee_member.id, 18000)
    await _make_finished_shift(
        db_session,
        verified_user.id,
        org.id,
        datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 12, 0, tzinfo=UTC),
    )
    await _create_adjustment(
        client,
        owner_headers,
        org.id,
        member_id=str(employee_member.id),
        amount_minor=15000,
        reason="Премия",
        occurred_at="2026-06-10T12:00:00Z",
    )
    resp = await client.get(
        f"/api/v1/organizations/{org.id}/payroll/export", headers=owner_headers
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    summary = wb["Сводка"]
    rows = list(summary.iter_rows(values_only=True))
    header = next(r for r in rows if r and r[0] == "Сотрудник")
    assert "Начисления/удержания, ₽" in header
    idx = header.index("Начисления/удержания, ₽")
    n_idx = header.index("К выплате, ₽")
    emp_row = next(r for r in rows if r and r[0] == "Test User")
    assert emp_row[idx] == 150.0  # 15000 коп.
    assert emp_row[n_idx] == 510.0  # 51000 коп.
