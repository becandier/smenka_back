# tests/test_payroll.py
"""Фича payroll: история ставок участников и расчёт зарплаты."""

import uuid
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from src.app.core.security import hash_password
from src.app.models.member_rate import OrganizationMemberRate, RateType
from src.app.models.organization import MemberRole, Organization, OrganizationMember
from src.app.models.shift import Pause, Shift, ShiftStatus
from src.app.models.shift_overtime_request import OvertimeRequestStatus, ShiftOvertimeRequest
from src.app.models.user import User
from src.app.models.work_location import WorkLocation

RATE_EFF_JAN = datetime(2026, 1, 1, 0, 0, 0, tzinfo=UTC)


async def _make_finished_shift(
    db_session: AsyncSession,
    user_id: uuid.UUID,
    org_id: uuid.UUID,
    started_at: datetime,
    finished_at: datetime,
    status: ShiftStatus = ShiftStatus.finished,
    work_location_id: uuid.UUID | None = None,
) -> Shift:
    shift = Shift(
        user_id=user_id,
        organization_id=org_id,
        started_at=started_at,
        finished_at=finished_at if status == ShiftStatus.finished else None,
        status=status,
        work_location_id=work_location_id,
    )
    db_session.add(shift)
    await db_session.commit()
    return shift


async def _make_work_location(
    db_session: AsyncSession,
    org_id: uuid.UUID,
    name: str = "Точка",
) -> WorkLocation:
    loc = WorkLocation(
        organization_id=org_id,
        name=name,
        latitude=55.75,
        longitude=37.62,
        radius_meters=100,
    )
    db_session.add(loc)
    await db_session.commit()
    return loc


async def _make_rate(
    db_session: AsyncSession,
    member_id: uuid.UUID,
    amount: int,
    rate_type: RateType = RateType.hourly,
    effective_from: datetime = RATE_EFF_JAN,
) -> OrganizationMemberRate:
    rate = OrganizationMemberRate(
        member_id=member_id,
        rate_amount_minor=amount,
        rate_type=rate_type,
        currency="RUB",
        effective_from=effective_from,
    )
    db_session.add(rate)
    await db_session.commit()
    return rate


@pytest.fixture
async def owner(db_session: AsyncSession) -> User:
    user = User(
        id=uuid.uuid4(),
        email="owner@example.com",
        password_hash=hash_password("Test1234"),
        name="Owner",
        is_verified=True,
    )
    db_session.add(user)
    await db_session.commit()
    return user


@pytest.fixture
async def owner_headers(owner: User, client: AsyncClient) -> dict[str, str]:
    resp = await client.post(
        "/api/v1/auth/login",
        json={"email": "owner@example.com", "password": "Test1234"},
    )
    token = resp.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
async def admin_user(db_session: AsyncSession) -> User:
    user = User(
        id=uuid.uuid4(),
        email="orgadmin@example.com",
        password_hash=hash_password("Test1234"),
        name="Org Admin",
        is_verified=True,
    )
    db_session.add(user)
    await db_session.commit()
    return user


@pytest.fixture
async def admin_headers(admin_user: User, client: AsyncClient) -> dict[str, str]:
    resp = await client.post(
        "/api/v1/auth/login",
        json={"email": "orgadmin@example.com", "password": "Test1234"},
    )
    token = resp.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
async def org(db_session: AsyncSession, owner: User) -> Organization:
    organization = Organization(name="Payroll Org", owner_id=owner.id)
    db_session.add(organization)
    await db_session.commit()
    return organization


@pytest.fixture
async def employee_member(
    db_session: AsyncSession,
    org: Organization,
    verified_user: User,
) -> OrganizationMember:
    """verified_user (conftest) как employee организации."""
    member = OrganizationMember(
        organization_id=org.id,
        user_id=verified_user.id,
        role=MemberRole.employee,
    )
    db_session.add(member)
    await db_session.commit()
    return member


@pytest.fixture
async def admin_member(
    db_session: AsyncSession,
    org: Organization,
    admin_user: User,
) -> OrganizationMember:
    member = OrganizationMember(
        organization_id=org.id,
        user_id=admin_user.id,
        role=MemberRole.admin,
    )
    db_session.add(member)
    await db_session.commit()
    return member


def _rates_url(org_id: Any, member_id: Any) -> str:
    return f"/api/v1/organizations/{org_id}/members/{member_id}/rates"


VALID_RATE_BODY = {
    "rate_amount_minor": 18000,
    "rate_type": "hourly",
    "effective_from": "2026-03-01T00:00:00Z",
    "note": "повышение",
}


class TestCreateRate:
    async def test_owner_creates_rate(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.post(
            _rates_url(org.id, employee_member.id),
            headers=owner_headers,
            json=VALID_RATE_BODY,
        )
        assert resp.status_code == 201
        data = resp.json()["data"]
        assert data["member_id"] == str(employee_member.id)
        assert data["rate_amount_minor"] == 18000
        assert data["rate_type"] == "hourly"
        assert data["currency"] == "RUB"
        assert data["effective_from"].startswith("2026-03-01T00:00:00")
        assert data["note"] == "повышение"
        assert data["id"]
        assert data["created_at"]

    async def test_admin_creates_rate(
        self,
        client: AsyncClient,
        admin_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
        admin_member: OrganizationMember,
    ) -> None:
        resp = await client.post(
            _rates_url(org.id, employee_member.id),
            headers=admin_headers,
            json={**VALID_RATE_BODY, "rate_type": "per_shift"},
        )
        assert resp.status_code == 201
        assert resp.json()["data"]["rate_type"] == "per_shift"

    async def test_employee_forbidden(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.post(
            _rates_url(org.id, employee_member.id),
            headers=auth_headers,
            json=VALID_RATE_BODY,
        )
        assert resp.status_code == 403
        assert resp.json()["error"]["code"] == "FORBIDDEN"

    async def test_duplicate_effective_from(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        first = await client.post(
            _rates_url(org.id, employee_member.id),
            headers=owner_headers,
            json=VALID_RATE_BODY,
        )
        assert first.status_code == 201
        second = await client.post(
            _rates_url(org.id, employee_member.id),
            headers=owner_headers,
            json={**VALID_RATE_BODY, "rate_amount_minor": 20000},
        )
        assert second.status_code == 409
        assert second.json()["error"]["code"] == "RATE_EFFECTIVE_FROM_TAKEN"

    async def test_non_positive_amount_validation(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.post(
            _rates_url(org.id, employee_member.id),
            headers=owner_headers,
            json={**VALID_RATE_BODY, "rate_amount_minor": 0},
        )
        assert resp.status_code == 422
        assert resp.json()["error"]["code"] == "VALIDATION_ERROR"

    async def test_unknown_rate_type_validation(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.post(
            _rates_url(org.id, employee_member.id),
            headers=owner_headers,
            json={**VALID_RATE_BODY, "rate_type": "monthly"},
        )
        assert resp.status_code == 422
        assert resp.json()["error"]["code"] == "VALIDATION_ERROR"

    async def test_member_not_found(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
    ) -> None:
        resp = await client.post(
            _rates_url(org.id, uuid.uuid4()),
            headers=owner_headers,
            json=VALID_RATE_BODY,
        )
        assert resp.status_code == 404
        assert resp.json()["error"]["code"] == "MEMBER_NOT_FOUND"

    async def test_member_of_other_org_not_found(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        owner: User,
        org: Organization,
        admin_user: User,
    ) -> None:
        """member_id чужой организации в пути этой org → 404."""
        other_org = Organization(name="Other Org", owner_id=owner.id)
        db_session.add(other_org)
        await db_session.flush()
        other_member = OrganizationMember(
            organization_id=other_org.id,
            user_id=admin_user.id,
            role=MemberRole.employee,
        )
        db_session.add(other_member)
        await db_session.commit()

        resp = await client.post(
            _rates_url(org.id, other_member.id),
            headers=owner_headers,
            json=VALID_RATE_BODY,
        )
        assert resp.status_code == 404
        assert resp.json()["error"]["code"] == "MEMBER_NOT_FOUND"

    async def test_org_not_found(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.post(
            _rates_url(uuid.uuid4(), employee_member.id),
            headers=owner_headers,
            json=VALID_RATE_BODY,
        )
        assert resp.status_code == 404
        assert resp.json()["error"]["code"] == "ORG_NOT_FOUND"


class TestRateHistory:
    async def test_list_sorted_desc(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(
            db_session,
            employee_member.id,
            15000,
            effective_from=datetime(2026, 1, 1, tzinfo=UTC),
        )
        await _make_rate(
            db_session,
            employee_member.id,
            18000,
            effective_from=datetime(2026, 3, 1, tzinfo=UTC),
        )
        await _make_rate(
            db_session,
            employee_member.id,
            16000,
            effective_from=datetime(2026, 2, 1, tzinfo=UTC),
        )
        resp = await client.get(
            _rates_url(org.id, employee_member.id),
            headers=owner_headers,
        )
        assert resp.status_code == 200
        items = resp.json()["data"]["items"]
        assert [i["rate_amount_minor"] for i in items] == [18000, 16000, 15000]

    async def test_employee_forbidden(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.get(
            _rates_url(org.id, employee_member.id),
            headers=auth_headers,
        )
        assert resp.status_code == 403


class TestUpdateRate:
    async def test_fix_amount(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        rate = await _make_rate(db_session, employee_member.id, 18000)
        resp = await client.patch(
            f"{_rates_url(org.id, employee_member.id)}/{rate.id}",
            headers=owner_headers,
            json={"rate_amount_minor": 18500, "note": "опечатка"},
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["rate_amount_minor"] == 18500
        assert data["note"] == "опечатка"
        assert data["rate_type"] == "hourly"

    async def test_move_effective_from_to_taken_date(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(
            db_session,
            employee_member.id,
            15000,
            effective_from=datetime(2026, 1, 1, tzinfo=UTC),
        )
        rate2 = await _make_rate(
            db_session,
            employee_member.id,
            18000,
            effective_from=datetime(2026, 3, 1, tzinfo=UTC),
        )
        resp = await client.patch(
            f"{_rates_url(org.id, employee_member.id)}/{rate2.id}",
            headers=owner_headers,
            json={"effective_from": "2026-01-01T00:00:00Z"},
        )
        assert resp.status_code == 409
        assert resp.json()["error"]["code"] == "RATE_EFFECTIVE_FROM_TAKEN"

    async def test_move_effective_from_to_free_date(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        rate = await _make_rate(db_session, employee_member.id, 18000)
        resp = await client.patch(
            f"{_rates_url(org.id, employee_member.id)}/{rate.id}",
            headers=owner_headers,
            json={"effective_from": "2026-03-02T00:00:00Z"},
        )
        assert resp.status_code == 200
        assert resp.json()["data"]["effective_from"].startswith("2026-03-02T00:00:00")

    async def test_rate_not_found(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.patch(
            f"{_rates_url(org.id, employee_member.id)}/{uuid.uuid4()}",
            headers=owner_headers,
            json={"rate_amount_minor": 18500},
        )
        assert resp.status_code == 404
        assert resp.json()["error"]["code"] == "RATE_NOT_FOUND"

    async def test_rate_of_other_member_not_found(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        employee_member: OrganizationMember,
        admin_member: OrganizationMember,
    ) -> None:
        rate = await _make_rate(db_session, admin_member.id, 18000)
        resp = await client.patch(
            f"{_rates_url(org.id, employee_member.id)}/{rate.id}",
            headers=owner_headers,
            json={"rate_amount_minor": 18500},
        )
        assert resp.status_code == 404
        assert resp.json()["error"]["code"] == "RATE_NOT_FOUND"


class TestDeleteRate:
    async def test_delete_then_404(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        rate = await _make_rate(db_session, employee_member.id, 18000)
        url = f"{_rates_url(org.id, employee_member.id)}/{rate.id}"
        resp = await client.delete(url, headers=owner_headers)
        assert resp.status_code == 200
        assert resp.json()["data"] == {"deleted": True}

        again = await client.delete(url, headers=owner_headers)
        assert again.status_code == 404
        assert again.json()["error"]["code"] == "RATE_NOT_FOUND"


class TestCurrentRateInMembers:
    async def test_owner_sees_latest_effective_rate(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(
            db_session,
            employee_member.id,
            15000,
            effective_from=datetime(2026, 1, 1, tzinfo=UTC),
        )
        await _make_rate(
            db_session,
            employee_member.id,
            18000,
            effective_from=datetime(2026, 3, 1, tzinfo=UTC),
        )
        # будущая ставка не действует
        await _make_rate(
            db_session,
            employee_member.id,
            99000,
            effective_from=datetime(2030, 1, 1, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/members",
            headers=owner_headers,
        )
        assert resp.status_code == 200
        items = resp.json()["data"]["items"]
        member_payload = next(i for i in items if i["id"] == str(employee_member.id))
        assert member_payload["current_rate"] is not None
        assert member_payload["current_rate"]["rate_amount_minor"] == 18000
        assert member_payload["current_rate"]["rate_type"] == "hourly"
        assert member_payload["current_rate"]["currency"] == "RUB"

    async def test_future_only_rates_give_null(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(
            db_session,
            employee_member.id,
            18000,
            effective_from=datetime(2030, 1, 1, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/members",
            headers=owner_headers,
        )
        items = resp.json()["data"]["items"]
        assert all(i["current_rate"] is None for i in items)

    async def test_no_rates_give_null(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/members",
            headers=owner_headers,
        )
        items = resp.json()["data"]["items"]
        assert all(i["current_rate"] is None for i in items)

    async def test_employee_never_sees_rates(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        employee_member: OrganizationMember,
        admin_member: OrganizationMember,
    ) -> None:
        """Ставки видят только owner/admin: для employee current_rate = null."""
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_rate(db_session, admin_member.id, 25000)
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/members",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        items = resp.json()["data"]["items"]
        assert len(items) == 2
        assert all(i["current_rate"] is None for i in items)

    async def test_admin_sees_rates(
        self,
        client: AsyncClient,
        admin_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        employee_member: OrganizationMember,
        admin_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/members",
            headers=admin_headers,
        )
        items = resp.json()["data"]["items"]
        member_payload = next(i for i in items if i["id"] == str(employee_member.id))
        assert member_payload["current_rate"]["rate_amount_minor"] == 18000


class TestPayrollReport:
    async def test_hourly_calculation_and_totals(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """180 ₽/час (18000 коп.), 2ч + 1ч → 54000 коп."""
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 3, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 3, 11, 0, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
            params={
                "date_from": "2026-06-01T00:00:00Z",
                "date_to": "2026-06-30T23:59:59Z",
            },
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["currency"] == "RUB"
        assert data["period"]["date_from"].startswith("2026-06-01T00:00:00")
        assert len(data["items"]) == 1
        item = data["items"][0]
        assert item["user_id"] == str(verified_user.id)
        assert item["worked_seconds"] == 10800
        assert item["shifts_count"] == 2
        assert item["gross_amount_minor"] == 54000
        assert item["unpaid_seconds"] == 0
        assert item["unpaid_shifts_count"] == 0
        assert item["has_missing_rate"] is False
        # Смены без графика: план = факт (work_schedules, R8) -> delta = 0.
        assert item["planned_seconds"] == 10800
        assert item["planned_amount_minor"] == 54000
        assert item["delta_amount_minor"] == 0
        totals = data["totals"]
        assert totals["worked_seconds"] == 10800
        assert totals["shifts_count"] == 2
        assert totals["gross_amount_minor"] == 54000
        assert totals["penalty_amount_minor"] == 0
        assert totals["penalties_count"] == 0
        assert totals["net_amount_minor"] == 54000
        assert totals["planned_seconds"] == 10800
        assert totals["planned_amount_minor"] == 54000
        assert totals["delta_amount_minor"] == 0
        assert totals["late_count"] == 0

    async def test_rate_change_applies_by_shift_start(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """Каждая смена считается по ставке, действовавшей на её started_at."""
        await _make_rate(
            db_session,
            employee_member.id,
            10000,
            effective_from=datetime(2026, 1, 1, tzinfo=UTC),
        )
        await _make_rate(
            db_session,
            employee_member.id,
            20000,
            effective_from=datetime(2026, 6, 2, tzinfo=UTC),
        )
        # 2ч по старой ставке + 1ч по новой
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 3, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 3, 11, 0, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
        )
        assert resp.status_code == 200
        item = resp.json()["data"]["items"][0]
        assert item["gross_amount_minor"] == 2 * 10000 + 1 * 20000

    async def test_per_shift_rate_ignores_duration(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(
            db_session,
            employee_member.id,
            50000,
            rate_type=RateType.per_shift,
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 3, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 3, 10, 30, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
        )
        item = resp.json()["data"]["items"][0]
        assert item["gross_amount_minor"] == 100000

    async def test_mixed_rate_types_in_history(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """rate_type меняется по истории: каждая смена — по типу своей ставки."""
        await _make_rate(
            db_session,
            employee_member.id,
            10000,
            rate_type=RateType.hourly,
            effective_from=datetime(2026, 1, 1, tzinfo=UTC),
        )
        await _make_rate(
            db_session,
            employee_member.id,
            30000,
            rate_type=RateType.per_shift,
            effective_from=datetime(2026, 6, 2, tzinfo=UTC),
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 11, 0, tzinfo=UTC),
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 3, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 3, 12, 0, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
        )
        item = resp.json()["data"]["items"][0]
        assert item["gross_amount_minor"] == 10000 + 30000

    async def test_shift_before_first_rate_is_unpaid(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(
            db_session,
            employee_member.id,
            18000,
            effective_from=datetime(2026, 6, 2, tzinfo=UTC),
        )
        # до первой ставки — неоплачиваемая
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 3, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 3, 11, 0, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
        )
        item = resp.json()["data"]["items"][0]
        assert item["worked_seconds"] == 10800
        assert item["shifts_count"] == 2
        assert item["gross_amount_minor"] == 18000
        assert item["unpaid_seconds"] == 7200
        assert item["unpaid_shifts_count"] == 1
        assert item["has_missing_rate"] is True

    async def test_rounding_half_up_once_on_employee_total(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """Две смены по 0.5 коп. дробной части: построчное округление дало бы
        10002, единое округление итога — 10001."""
        await _make_rate(db_session, employee_member.id, 10001)
        for day in (1, 3):
            await _make_finished_shift(
                db_session,
                verified_user.id,
                org.id,
                datetime(2026, 6, day, 10, 0, tzinfo=UTC),
                datetime(2026, 6, day, 10, 30, tzinfo=UTC),
            )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
        )
        item = resp.json()["data"]["items"][0]
        assert item["gross_amount_minor"] == 10001

    async def test_pauses_reduce_paid_time(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """worked_seconds — за вычетом пауз (calculate_worked_seconds)."""
        await _make_rate(db_session, employee_member.id, 36000)
        shift = await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        pause = Pause(
            shift_id=shift.id,
            started_at=datetime(2026, 6, 1, 10, 30, tzinfo=UTC),
            finished_at=datetime(2026, 6, 1, 11, 0, tzinfo=UTC),
        )
        db_session.add(pause)
        await db_session.commit()

        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
        )
        item = resp.json()["data"]["items"][0]
        assert item["worked_seconds"] == 5400
        assert item["gross_amount_minor"] == 54000

    async def test_only_finished_shifts_counted(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
            status=ShiftStatus.active,
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
        )
        assert resp.json()["data"]["items"] == []

    async def test_period_boundaries_inclusive(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        boundary = datetime(2026, 6, 5, 10, 0, tzinfo=UTC)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            boundary,
            datetime(2026, 6, 5, 11, 0, tzinfo=UTC),
        )
        # вне окна
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 5, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 5, 1, 11, 0, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
            params={
                "date_from": "2026-06-01T00:00:00Z",
                "date_to": boundary.isoformat().replace("+00:00", "Z"),
            },
        )
        item = resp.json()["data"]["items"][0]
        assert item["shifts_count"] == 1

    async def test_removed_member_shifts_become_unpaid(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """Исключение участника каскадом удаляет ставки → его смены unpaid."""
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        await db_session.delete(employee_member)
        await db_session.commit()

        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
        )
        items = resp.json()["data"]["items"]
        assert len(items) == 1
        assert items[0]["user_name"] == "Test User"
        assert items[0]["gross_amount_minor"] == 0
        assert items[0]["has_missing_rate"] is True

    async def test_employee_forbidden(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=auth_headers,
        )
        assert resp.status_code == 403
        assert resp.json()["error"]["code"] == "FORBIDDEN"

    async def test_org_not_found(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
    ) -> None:
        resp = await client.get(
            f"/api/v1/organizations/{uuid.uuid4()}/payroll",
            headers=owner_headers,
        )
        assert resp.status_code == 404
        assert resp.json()["error"]["code"] == "ORG_NOT_FOUND"

    async def test_invalid_date_range(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
    ) -> None:
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
            params={
                "date_from": "2026-06-05T00:00:00Z",
                "date_to": "2026-06-01T00:00:00Z",
            },
        )
        assert resp.status_code == 400
        assert resp.json()["error"]["code"] == "INVALID_DATE_RANGE"


class TestPayrollWorkSchedules:
    """R8 (work_schedules): план против факта, переработка, опоздания в payroll."""

    async def test_planned_from_schedule_window_differs_from_actual(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """180 ₽/час, план 9ч (09:00-18:00), факт — 8ч (пришёл на час раньше, ушёл раньше)."""
        await _make_rate(db_session, employee_member.id, 18000)
        shift = Shift(
            user_id=verified_user.id,
            organization_id=org.id,
            started_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            finished_at=datetime(2026, 6, 1, 17, 0, tzinfo=UTC),  # 8ч факт
            status=ShiftStatus.finished,
            scheduled_start_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            scheduled_end_at=datetime(2026, 6, 1, 18, 0, tzinfo=UTC),  # 9ч план
        )
        db_session.add(shift)
        await db_session.commit()

        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
            params={"date_from": "2026-06-01T00:00:00Z", "date_to": "2026-06-30T23:59:59Z"},
        )
        item = resp.json()["data"]["items"][0]
        assert item["worked_seconds"] == 8 * 3600
        assert item["gross_amount_minor"] == 8 * 18000
        assert item["planned_seconds"] == 9 * 3600
        assert item["planned_amount_minor"] == 9 * 18000
        assert item["delta_amount_minor"] == 8 * 18000 - 9 * 18000
        assert item["delta_amount_minor"] < 0  # недозаработал

    async def test_shift_without_schedule_plan_equals_fact(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 13, 0, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
            params={"date_from": "2026-06-01T00:00:00Z", "date_to": "2026-06-30T23:59:59Z"},
        )
        item = resp.json()["data"]["items"][0]
        assert item["planned_seconds"] == item["worked_seconds"]
        assert item["planned_amount_minor"] == item["gross_amount_minor"]
        assert item["delta_amount_minor"] == 0

    async def test_shift_without_rate_excluded_from_planned_amount(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """Нет действующей ставки -> gross=0 и planned_amount_minor=0, delta=0."""
        shift = Shift(
            user_id=verified_user.id,
            organization_id=org.id,
            started_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            finished_at=datetime(2026, 6, 1, 17, 0, tzinfo=UTC),
            status=ShiftStatus.finished,
            scheduled_start_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            scheduled_end_at=datetime(2026, 6, 1, 18, 0, tzinfo=UTC),
        )
        db_session.add(shift)
        await db_session.commit()

        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
            params={"date_from": "2026-06-01T00:00:00Z", "date_to": "2026-06-30T23:59:59Z"},
        )
        item = resp.json()["data"]["items"][0]
        assert item["has_missing_rate"] is True
        assert item["gross_amount_minor"] == 0
        assert item["planned_amount_minor"] == 0
        assert item["delta_amount_minor"] == 0
        assert item["planned_seconds"] == 9 * 3600  # время само по себе не зависит от ставки

    async def test_per_shift_rate_plan_equals_fact_regardless_of_window(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 300000, rate_type=RateType.per_shift)
        shift = Shift(
            user_id=verified_user.id,
            organization_id=org.id,
            started_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            finished_at=datetime(2026, 6, 1, 17, 0, tzinfo=UTC),  # 8ч факт
            status=ShiftStatus.finished,
            scheduled_start_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            scheduled_end_at=datetime(2026, 6, 1, 18, 0, tzinfo=UTC),  # 9ч план
        )
        db_session.add(shift)
        await db_session.commit()

        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
            params={"date_from": "2026-06-01T00:00:00Z", "date_to": "2026-06-30T23:59:59Z"},
        )
        item = resp.json()["data"]["items"][0]
        assert item["gross_amount_minor"] == 300000
        assert item["planned_amount_minor"] == 300000
        assert item["delta_amount_minor"] == 0

    async def test_approved_overtime_increases_gross_for_hourly(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)  # 180 ₽/час
        shift = Shift(
            user_id=verified_user.id,
            organization_id=org.id,
            started_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            finished_at=datetime(2026, 6, 1, 18, 0, tzinfo=UTC),  # 9ч факт == план
            status=ShiftStatus.finished,
            scheduled_start_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            scheduled_end_at=datetime(2026, 6, 1, 18, 0, tzinfo=UTC),
        )
        db_session.add(shift)
        await db_session.flush()
        overtime = ShiftOvertimeRequest(
            shift_id=shift.id,
            minutes=30,
            comment="Задержался",
            status=OvertimeRequestStatus.approved,
        )
        db_session.add(overtime)
        await db_session.commit()

        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
            params={"date_from": "2026-06-01T00:00:00Z", "date_to": "2026-06-30T23:59:59Z"},
        )
        item = resp.json()["data"]["items"][0]
        assert item["overtime_seconds"] == 30 * 60
        # 9ч факт + 30 мин переработки, оплачено по часовой ставке.
        assert item["gross_amount_minor"] == round((9 * 3600 + 30 * 60) * 18000 / 3600)
        assert item["planned_amount_minor"] == 9 * 18000  # план не включает переработку
        assert item["delta_amount_minor"] > 0  # переработал -> заработал больше плана

    async def test_pending_overtime_does_not_affect_gross(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        shift = Shift(
            user_id=verified_user.id,
            organization_id=org.id,
            started_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            finished_at=datetime(2026, 6, 1, 18, 0, tzinfo=UTC),
            status=ShiftStatus.finished,
            scheduled_start_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            scheduled_end_at=datetime(2026, 6, 1, 18, 0, tzinfo=UTC),
        )
        db_session.add(shift)
        await db_session.flush()
        db_session.add(
            ShiftOvertimeRequest(
                shift_id=shift.id,
                minutes=30,
                comment="На согласовании",
                status=OvertimeRequestStatus.pending,
            )
        )
        await db_session.commit()

        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
            params={"date_from": "2026-06-01T00:00:00Z", "date_to": "2026-06-30T23:59:59Z"},
        )
        item = resp.json()["data"]["items"][0]
        assert item["overtime_seconds"] == 0
        assert item["gross_amount_minor"] == 9 * 18000

    async def test_late_count_and_seconds_with_tolerance(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        from src.app.models.organization_settings import OrganizationSettings

        db_session.add(OrganizationSettings(organization_id=org.id, late_tolerance_minutes=10))
        await _make_rate(db_session, employee_member.id, 18000)
        shift = Shift(
            user_id=verified_user.id,
            organization_id=org.id,
            started_at=datetime(2026, 6, 1, 9, 20, tzinfo=UTC),  # опоздал на 20 мин
            finished_at=datetime(2026, 6, 1, 18, 0, tzinfo=UTC),
            status=ShiftStatus.finished,
            scheduled_start_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
            scheduled_end_at=datetime(2026, 6, 1, 18, 0, tzinfo=UTC),
        )
        db_session.add(shift)
        await db_session.commit()

        resp = await client.get(
            f"/api/v1/organizations/{org.id}/payroll",
            headers=owner_headers,
            params={"date_from": "2026-06-01T00:00:00Z", "date_to": "2026-06-30T23:59:59Z"},
        )
        item = resp.json()["data"]["items"][0]
        assert item["late_count"] == 1
        # 20 мин опоздания - 10 мин допуска = 10 мин = 600 сек
        assert item["late_seconds_total"] == 600


class TestMyEarnings:
    async def test_employee_gets_own_earnings(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        admin_user: User,
        employee_member: OrganizationMember,
        admin_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_rate(db_session, admin_member.id, 99000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        # чужая смена не учитывается
        await _make_finished_shift(
            db_session,
            admin_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 18, 0, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/my-earnings",
            headers=auth_headers,
            params={
                "date_from": "2026-06-01T00:00:00Z",
                "date_to": "2026-06-30T00:00:00Z",
            },
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["currency"] == "RUB"
        assert data["worked_seconds"] == 7200
        assert data["shifts_count"] == 1
        assert data["gross_amount_minor"] == 36000
        assert data["has_missing_rate"] is False
        assert data["current_rate"]["rate_amount_minor"] == 18000
        assert data["period"]["date_from"].startswith("2026-06-01T00:00:00")

    async def test_no_rates_zero_gross_missing_flag(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/my-earnings",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["gross_amount_minor"] == 0
        assert data["worked_seconds"] == 7200
        assert data["has_missing_rate"] is True
        assert data["current_rate"] is None

    async def test_owner_forbidden(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
    ) -> None:
        """Owner — не member (ADR-001): в my-earnings не участвует."""
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/my-earnings",
            headers=owner_headers,
        )
        assert resp.status_code == 403
        assert resp.json()["error"]["code"] == "FORBIDDEN"

    async def test_non_member_forbidden(
        self,
        client: AsyncClient,
        admin_headers: dict[str, Any],
        org: Organization,
    ) -> None:
        """Пользователь без членства (admin_member не создан) → 403."""
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/my-earnings",
            headers=admin_headers,
        )
        assert resp.status_code == 403

    async def test_org_not_found(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
    ) -> None:
        resp = await client.get(
            f"/api/v1/organizations/{uuid.uuid4()}/my-earnings",
            headers=auth_headers,
        )
        assert resp.status_code == 404
        assert resp.json()["error"]["code"] == "ORG_NOT_FOUND"

    async def test_invalid_date_range(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.get(
            f"/api/v1/organizations/{org.id}/my-earnings",
            headers=auth_headers,
            params={
                "date_from": "2026-06-05T00:00:00Z",
                "date_to": "2026-06-01T00:00:00Z",
            },
        )
        assert resp.status_code == 400
        assert resp.json()["error"]["code"] == "INVALID_DATE_RANGE"


def _payroll_url(org_id: Any) -> str:
    return f"/api/v1/organizations/{org_id}/payroll"


def _export_url(org_id: Any) -> str:
    return f"/api/v1/organizations/{org_id}/payroll/export"


async def _make_employee(
    db_session: AsyncSession,
    org: Organization,
    email: str,
    name: str,
) -> tuple[User, OrganizationMember]:
    user = User(
        id=uuid.uuid4(),
        email=email,
        password_hash=hash_password("Test1234"),
        name=name,
        is_verified=True,
    )
    db_session.add(user)
    await db_session.flush()
    member = OrganizationMember(
        organization_id=org.id,
        user_id=user.id,
        role=MemberRole.employee,
    )
    db_session.add(member)
    await db_session.commit()
    return user, member


class TestPayrollDetailed:
    async def test_granularity_none_keeps_legacy_shape(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """granularity=none: ни breakdown, ни granularity/tz в ответе."""
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        resp = await client.get(_payroll_url(org.id), headers=owner_headers)
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert "granularity" not in data
        assert "tz" not in data
        assert "breakdown" not in data["items"][0]

    async def test_day_breakdown_buckets_and_echo(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 3, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 3, 11, 0, tzinfo=UTC),
        )
        resp = await client.get(
            _payroll_url(org.id), headers=owner_headers, params={"granularity": "day"}
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["granularity"] == "day"
        assert data["tz"] == "UTC"
        item = data["items"][0]
        breakdown = item["breakdown"]
        assert [b["bucket_start"] for b in breakdown] == ["2026-06-01", "2026-06-03"]
        assert breakdown[0]["worked_seconds"] == 7200
        assert breakdown[0]["gross_amount_minor"] == 36000
        assert breakdown[0]["shifts_count"] == 1
        assert breakdown[1]["gross_amount_minor"] == 18000
        # суммы корзин сходятся в итог сотрудника и totals
        assert item["gross_amount_minor"] == 54000
        assert item["worked_seconds"] == 10800
        assert sum(b["gross_amount_minor"] for b in breakdown) == item["gross_amount_minor"]
        assert data["totals"]["gross_amount_minor"] == 54000

    async def test_day_rounding_is_atomic_per_day(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """Посуточное округление: 2×0.5коп. → 5001+5001=10002 (none даёт 10001)."""
        await _make_rate(db_session, employee_member.id, 10001)
        for day in (1, 3):
            await _make_finished_shift(
                db_session,
                verified_user.id,
                org.id,
                datetime(2026, 6, day, 10, 0, tzinfo=UTC),
                datetime(2026, 6, day, 10, 30, tzinfo=UTC),
            )
        detailed = await client.get(
            _payroll_url(org.id), headers=owner_headers, params={"granularity": "day"}
        )
        item = detailed.json()["data"]["items"][0]
        assert [b["gross_amount_minor"] for b in item["breakdown"]] == [5001, 5001]
        assert item["gross_amount_minor"] == 10002
        assert detailed.json()["data"]["totals"]["gross_amount_minor"] == 10002
        # режим none сохраняет единичное округление (обратная совместимость)
        legacy = await client.get(_payroll_url(org.id), headers=owner_headers)
        assert legacy.json()["data"]["items"][0]["gross_amount_minor"] == 10001

    async def test_week_granularity_groups_iso_week(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        for day in (1, 3, 8):  # 01 и 03 — одна неделя (пн 01), 08 — следующая
            await _make_finished_shift(
                db_session,
                verified_user.id,
                org.id,
                datetime(2026, 6, day, 10, 0, tzinfo=UTC),
                datetime(2026, 6, day, 11, 0, tzinfo=UTC),
            )
        resp = await client.get(
            _payroll_url(org.id), headers=owner_headers, params={"granularity": "week"}
        )
        breakdown = resp.json()["data"]["items"][0]["breakdown"]
        assert [b["bucket_start"] for b in breakdown] == ["2026-06-01", "2026-06-08"]
        assert breakdown[0]["shifts_count"] == 2
        assert breakdown[0]["gross_amount_minor"] == 36000
        assert breakdown[1]["shifts_count"] == 1

    async def test_month_granularity_buckets(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 10, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 10, 11, 0, tzinfo=UTC),
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 7, 5, 10, 0, tzinfo=UTC),
            datetime(2026, 7, 5, 11, 0, tzinfo=UTC),
        )
        resp = await client.get(
            _payroll_url(org.id), headers=owner_headers, params={"granularity": "month"}
        )
        breakdown = resp.json()["data"]["items"][0]["breakdown"]
        assert [b["bucket_start"] for b in breakdown] == ["2026-06-01", "2026-07-01"]

    async def test_tz_shifts_bucket_to_local_day(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        """Смена 21:30 UTC (00:30 МСК) попадает в день по локальной таймзоне."""
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 21, 30, tzinfo=UTC),
            datetime(2026, 6, 1, 22, 30, tzinfo=UTC),
        )
        utc = await client.get(
            _payroll_url(org.id), headers=owner_headers, params={"granularity": "day"}
        )
        assert utc.json()["data"]["items"][0]["breakdown"][0]["bucket_start"] == "2026-06-01"

        msk = await client.get(
            _payroll_url(org.id),
            headers=owner_headers,
            params={"granularity": "day", "tz": "Europe/Moscow"},
        )
        data = msk.json()["data"]
        assert data["tz"] == "Europe/Moscow"
        assert data["items"][0]["breakdown"][0]["bucket_start"] == "2026-06-02"

    async def test_invalid_tz_422(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
    ) -> None:
        resp = await client.get(
            _payroll_url(org.id),
            headers=owner_headers,
            params={"granularity": "day", "tz": "Mars/Phobos"},
        )
        assert resp.status_code == 422
        assert resp.json()["error"]["code"] == "VALIDATION_ERROR"

    async def test_invalid_granularity_422(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
    ) -> None:
        resp = await client.get(
            _payroll_url(org.id), headers=owner_headers, params={"granularity": "year"}
        )
        assert resp.status_code == 422
        assert resp.json()["error"]["code"] == "VALIDATION_ERROR"

    async def test_user_ids_filter(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        emp2_user, emp2_member = await _make_employee(
            db_session, org, "emp2@example.com", "Employee Two"
        )
        await _make_rate(db_session, emp2_member.id, 18000)
        for user in (verified_user, emp2_user):
            await _make_finished_shift(
                db_session,
                user.id,
                org.id,
                datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
                datetime(2026, 6, 1, 11, 0, tzinfo=UTC),
            )
        only_first = await client.get(
            _payroll_url(org.id),
            headers=owner_headers,
            params={"user_ids": str(verified_user.id)},
        )
        items = only_first.json()["data"]["items"]
        assert [i["user_id"] for i in items] == [str(verified_user.id)]

        csv_both = await client.get(
            _payroll_url(org.id),
            headers=owner_headers,
            params={"user_ids": f"{verified_user.id},{emp2_user.id}"},
        )
        assert len(csv_both.json()["data"]["items"]) == 2

    async def test_user_ids_bad_uuid_422(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
    ) -> None:
        resp = await client.get(
            _payroll_url(org.id), headers=owner_headers, params={"user_ids": "not-a-uuid"}
        )
        assert resp.status_code == 422
        assert resp.json()["error"]["code"] == "VALIDATION_ERROR"

    async def test_location_ids_filter_including_no_location(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        loc1 = await _make_work_location(db_session, org.id, "Точка 1")
        loc2 = await _make_work_location(db_session, org.id, "Точка 2")
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
            work_location_id=loc1.id,
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 2, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 2, 11, 0, tzinfo=UTC),
            work_location_id=loc2.id,
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 3, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 3, 10, 30, tzinfo=UTC),
        )

        async def _worked(params: dict[str, Any]) -> int:
            resp = await client.get(_payroll_url(org.id), headers=owner_headers, params=params)
            items = resp.json()["data"]["items"]
            return items[0]["worked_seconds"] if items else 0

        assert await _worked({"location_ids": str(loc1.id)}) == 7200
        assert await _worked({"location_ids": "none"}) == 1800
        assert await _worked({"location_ids": f"{loc1.id},none"}) == 9000

    async def test_location_foreign_org_422(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        owner: User,
    ) -> None:
        other_org = Organization(name="Other Org", owner_id=owner.id)
        db_session.add(other_org)
        await db_session.flush()
        foreign = await _make_work_location(db_session, other_org.id, "Чужая точка")
        resp = await client.get(
            _payroll_url(org.id),
            headers=owner_headers,
            params={"location_ids": str(foreign.id)},
        )
        assert resp.status_code == 422
        assert resp.json()["error"]["code"] == "VALIDATION_ERROR"

    async def test_only_missing_rate_filter(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)  # оплачиваемый
        emp2_user, _ = await _make_employee(db_session, org, "emp2@example.com", "No Rate")
        for user in (verified_user, emp2_user):
            await _make_finished_shift(
                db_session,
                user.id,
                org.id,
                datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
                datetime(2026, 6, 1, 11, 0, tzinfo=UTC),
            )
        resp = await client.get(
            _payroll_url(org.id),
            headers=owner_headers,
            params={"only_missing_rate": "true"},
        )
        items = resp.json()["data"]["items"]
        assert [i["user_id"] for i in items] == [str(emp2_user.id)]
        assert items[0]["has_missing_rate"] is True


class TestPayrollExport:
    async def test_export_valid_xlsx_with_sheets(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 3, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 3, 11, 0, tzinfo=UTC),
        )
        resp = await client.get(_export_url(org.id), headers=owner_headers)
        assert resp.status_code == 200
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment; filename=" in resp.headers["content-disposition"]

        wb = load_workbook(BytesIO(resp.content))
        assert wb.sheetnames == ["Сводка", "Детализация"]

        summary = wb["Сводка"]
        rows = list(summary.iter_rows(values_only=True))
        total_row = next(r for r in rows if r[0] == "ИТОГО")
        assert total_row[1] == 3.0  # часы (2ч + 1ч)
        assert total_row[2] == 2  # смены
        assert total_row[3] == 540.0  # рубли (18000 коп/ч × 3ч)

        detail = wb["Детализация"]
        detail_rows = list(detail.iter_rows(min_row=2, values_only=True))
        assert [r[1] for r in detail_rows] == ["2026-06-01", "2026-06-03"]
        # деньги детализации суммируются в сводку (число в рублях)
        assert sum(r[4] for r in detail_rows) == total_row[3]

    async def test_export_unsupported_format_422(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        org: Organization,
    ) -> None:
        resp = await client.get(
            _export_url(org.id), headers=owner_headers, params={"format": "csv"}
        )
        assert resp.status_code == 422
        assert resp.json()["error"]["code"] == "VALIDATION_ERROR"

    async def test_export_employee_forbidden(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        org: Organization,
        employee_member: OrganizationMember,
    ) -> None:
        resp = await client.get(_export_url(org.id), headers=auth_headers)
        assert resp.status_code == 403
        assert resp.json()["error"]["code"] == "FORBIDDEN"

    async def test_export_org_not_found(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
    ) -> None:
        resp = await client.get(_export_url(uuid.uuid4()), headers=owner_headers)
        assert resp.status_code == 404
        assert resp.json()["error"]["code"] == "ORG_NOT_FOUND"

    async def test_export_filename_contains_period_dates(
        self,
        client: AsyncClient,
        owner_headers: dict[str, Any],
        db_session: AsyncSession,
        org: Organization,
        verified_user: User,
        employee_member: OrganizationMember,
    ) -> None:
        await _make_rate(db_session, employee_member.id, 18000)
        await _make_finished_shift(
            db_session,
            verified_user.id,
            org.id,
            datetime(2026, 6, 1, 10, 0, tzinfo=UTC),
            datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        )
        resp = await client.get(
            _export_url(org.id),
            headers=owner_headers,
            params={
                "date_from": "2026-06-01T00:00:00Z",
                "date_to": "2026-06-30T23:59:59Z",
            },
        )
        assert resp.status_code == 200
        disposition = resp.headers["content-disposition"]
        assert "2026-06-01" in disposition
        assert "2026-06-30" in disposition
